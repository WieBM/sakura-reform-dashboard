"""
Sakura Reform Dashboard — Excel Source Data Inspector
Usage:
  python scripts/excel_inspect.py                    # summary of all Excel files
  python scripts/excel_inspect.py --file 社員名簿     # first 10 rows of that file
"""

import argparse
import os
import sys
import openpyxl
sys.stdout.reconfigure(encoding="utf-8")

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "HandsOn_資料")

FILES = {
    "社員名簿":       ("社員名簿.xlsx",          "社員"),
    "顧客管理台帳":    ("顧客管理台帳.xlsx",       "顧客"),
    "案件管理表":      ("案件管理表.xlsx",         "案件"),
    "見積明細一覧":    ("見積明細一覧.xlsx",        "見積"),
    "請求入金管理表":  ("請求・入金管理表.xlsx",    "請求"),
    "工事サービス単価": ("工事サービス標準単価表.xlsx", "工事サービス"),
}

HEADER_ROW = 4  # db.js skips first 4 rows (slice(4))


def summary():
    for key, (filename, sheet) in FILES.items():
        path = os.path.join(DATA_DIR, filename)
        if not os.path.exists(path):
            print(f"  {key:<20} ⚠  file not found")
            continue
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet]
        # Count non-empty rows after header
        data_rows = sum(
            1 for i, row in enumerate(ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True))
            if any(c is not None for c in row)
        )
        headers = [c.value for c in next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW))]
        headers = [h for h in headers if h is not None]
        wb.close()
        print(f"  {key:<20} {data_rows:>4} rows   cols: {', '.join(str(h) for h in headers)}")


def show_file(key, limit=10):
    if key not in FILES:
        print(f"Unknown file key: {key}")
        print(f"Valid keys: {', '.join(FILES.keys())}")
        return
    filename, sheet = FILES[key]
    path = os.path.join(DATA_DIR, filename)
    if not os.path.exists(path):
        print(f"File not found: {path}")
        return

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows = list(ws.iter_rows(min_row=HEADER_ROW, values_only=True))
    headers = [str(h) if h is not None else "" for h in rows[0]]
    data = [r for r in rows[1:] if any(c is not None for c in r)][:limit]

    print(f"\n{filename} / sheet={sheet} (first {limit} data rows)")
    col_w = [max(len(str(h)), max((len(str(r[i])) if r[i] is not None else 0) for r in data) if data else 0)
             for i, h in enumerate(headers)]
    sep = "+-" + "-+-".join("-" * w for w in col_w) + "-+"
    fmt = "| " + " | ".join(f"{{:<{w}}}" for w in col_w) + " |"
    print(sep)
    print(fmt.format(*headers))
    print(sep)
    for r in data:
        print(fmt.format(*[str(v) if v is not None else "" for v in r]))
    print(sep)
    wb.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", "-f", help="File key to preview (e.g. 社員名簿)")
    args = parser.parse_args()

    if args.file:
        show_file(args.file)
    else:
        print("=== HandsOn_資料 -- Excel file summary ===")
        summary()
        print("\nOption: --file <key>  to preview rows")
